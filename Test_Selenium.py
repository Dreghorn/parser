from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# URL страницы товара
url = 'https://rozetka.com.ua/ua/apple-iphone-15-128gb-black/p395460480/'

# Запуск браузера
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # Без графического интерфейса
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get(url)

wait = WebDriverWait(driver, 10)  # Ожидание элементов

try:
    # Имя продукта
    product_name = wait.until(EC.presence_of_element_located((By.TAG_NAME, "h1"))).text.strip()

    # Цвет
    color = "Не найден"
    color_elements = driver.find_elements(By.CSS_SELECTOR, 'p.text-base.mb-2')
    for element in color_elements:
        if "Колір" in element.text:
            color = element.find_element(By.CSS_SELECTOR, "span.bold").text.strip()
            break

    # Встроенная память
    storage = "Не найден"
    for element in color_elements:
        if "Вбудована пам'ять" in element.text:
            storage = element.find_element(By.CSS_SELECTOR, "span.bold").text.strip()
            break

    # Продавец (логотип)
    try:
        seller_element = driver.find_element(By.CSS_SELECTOR, "span.seller-logo img")
        seller = seller_element.get_attribute("src")
    except:
        seller = "Продавец не найден"

    # Цены
    try:
        price_promo = driver.find_element(By.CSS_SELECTOR, "p.product-price__big.product-price__big-color-red").text.strip()
    except:
        price_promo = "Нет акционной цены"

    try:
        price_normal = driver.find_element(By.CSS_SELECTOR, "p.product-price__small").text.strip()
    except:
        price_normal = "Нет обычной цены"

    # Код товара
    try:
        # Выполним JavaScript для получения текста кода товара
        product_code = driver.execute_script("""
            let codeElement = document.querySelector("span.ms-auto.color-black-60");
            if (codeElement) {
                return codeElement.textContent.trim().replace('Код: ', '');
            } else {
                return 'Не найден';
            }
        """)
    except Exception as e:
        print(f"Ошибка при извлечении кода товара: {e}")

    # Количество отзывов
    try:
        reviews_count = driver.find_element(By.CSS_SELECTOR, "span.tabs__link-text").text.strip()
    except:
        reviews_count = "Нет отзывов"

    # Ссылки на изображения
    img_urls = []
    try:
        images = driver.find_elements(By.CSS_SELECTOR, "ul.simple-slider__list img.image")
        img_urls = [img.get_attribute("src") for img in images if img.get_attribute("src")]
    except:
        img_urls = []

    # Функция для парсинга характеристик
    def get_characteristics():
        characteristics = {}

        # Переход на вкладку характеристик
        driver.get(url + "characteristics/")

        try:
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "h1")))  # Ожидаем загрузку

            sections = driver.find_elements(By.CSS_SELECTOR, "section.group")
            for section in sections:
                try:
                    category = section.find_element(By.CSS_SELECTOR, "h3.sub-heading").text.strip()
                    characteristics[category] = {}

                    items = section.find_elements(By.CSS_SELECTOR, "div.item")
                    for item in items:
                        try:
                            label = item.find_element(By.CSS_SELECTOR, "dt.label").text.strip()
                            values = [v.text.strip() for v in item.find_elements(By.CSS_SELECTOR, "span") if v.text.strip()]
                            characteristics[category][label] = ", ".join(values)
                        except:
                            continue
                except:
                    continue
        except:
            print("❌ Не удалось загрузить характеристики.")

        return characteristics
    # Получаем серию, диагональ и разрешение с главной страницы
    series = "Не найдено"
    screen_size = "Не найдено"
    resolution = "Не найдено"

    try:
        series_element = driver.find_element(By.XPATH, "//span[contains(text(), 'Серія')]")
        series = series_element.find_element(By.XPATH, "../following-sibling::dd").text.strip()
    except:
        pass

    try:
        screen_size_element = driver.find_element(By.XPATH, "//span[contains(text(), 'Діагональ екрана')]")
        screen_size = screen_size_element.find_element(By.XPATH, "../following-sibling::dd").text.strip()
    except:
        pass

    try:
        resolution_element = driver.find_element(By.XPATH, "//span[contains(text(), 'Роздільна здатність дисплея')]")
        resolution = resolution_element.find_element(By.XPATH, "../following-sibling::dd").text.strip()
    except:
        pass

    # Получаем характеристики
    specs = get_characteristics()

    # Вывод данных
    print(f'Полное название товара: {product_name}')
    print(f'Цвет: {color}')
    print(f'Объем памяти: {storage}')
    print(f'Продавец (логотип): {seller}')
    print(f'Цена обычная: {price_normal}')
    print(f'Цена акционная: {price_promo}')
    print('Ссылки на фото товара:')
    for link in img_urls:
        print(f'  {link}')
    print(f'Код товара: {product_code}')
    print(f'Количество отзывов: {reviews_count}')
    print(f'Серия: {series}')
    print(f'Диагональ экрана: {screen_size}')
    print(f'Разрешение дисплея: {resolution}')
    print(f'Характеристики: {specs}')

    # Запись в Excel
    wb = load_workbook("template.xlsx")
    ws = wb.active

    ws["A2"] = product_name
    ws["B2"] = color
    ws["C2"] = storage
    ws["D2"] = seller
    ws["E2"] = price_normal
    ws["F2"] = price_promo
    ws["G2"] = product_code
    ws["H2"] = reviews_count
    ws["I2"] = series
    ws["J2"] = screen_size
    ws["K2"] = resolution

    for i, url in enumerate(img_urls, start=2):
        ws[f"L{i}"] = url

    # Запись характеристик на отдельный лист
    ws_specs = wb.create_sheet("Характеристики")
    row = 1
    for category, items in specs.items():
        ws_specs[f"A{row}"] = category  # Категория
        row += 1
        for key, value in items.items():
            ws_specs[f"B{row}"] = key  # Название характеристики
            ws_specs[f"C{row}"] = value  # Значения характеристики
            row += 1

    wb.save("result.xlsx")
    print("✅ Данные успешно сохранены в result.xlsx")

except Exception as e:
    print("❌ Ошибка:", e)

finally:
    driver.quit()  # Закрываем браузер
