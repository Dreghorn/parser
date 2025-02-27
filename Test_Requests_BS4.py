import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# URL страницы товара
url = 'https://rozetka.com.ua/apple-iphone-15-128gb-black/p395460480/'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

response = requests.get(url, headers=headers)

if response.status_code == 200:
    soup = BeautifulSoup(response.text, 'html.parser')

    # Извлекаем данные
    # Имя продукта
    product_name = soup.find('h1', class_='title__font').text.strip()
    
    # Цвет ищем спомощью IF так как некоторые элементы имеют один class и HTML тег
    color_element = soup.find('p', class_='text-base mb-2')
    color = color_element.find('span', class_='bold').text.strip() if color_element and "Колір" in color_element.text else "Не найден"
    
    # Встроенная память, ищем так же как и цвет
    memory_element = soup.find_all('p', class_='text-base mb-2')
    storage = "Не найден"
    for item in memory_element:
        if "Вбудована пам'ять" in item.text:
            storage = item.find('span', class_='bold').text.strip()
            
    # Продавец
    seller_element = soup.select_one("span.seller-logo img")

    if seller_element and 'src' in seller_element.attrs:
        seller = seller_element["src"]  # Берем ссылку на изображение
    else:
        seller = "Продавец не найден"
    
    # Акционная цена 
    price_promo = soup.find('p', class_='product-price__big product-price__big-color-red').text.strip()
    
    # Обычная цена
    price_normal = soup.find('p', class_='product-price__small').text.strip()
    
    # Код продукта
    product_code = soup.find('span', class_='ms-auto color-black-60').text.strip()
    
    # Количество отзывов
    reviews_count = soup.find('span', class_='tabs__link-text').text.strip()
    
    
    image_section = soup.find('ul', class_='simple-slider__list')

    if image_section:
        # Поиск всех изображений с классом 'image'
        images = image_section.find_all('img', class_='image')
        
        # Список ссылок на изображения
        img_urls = [img['src'] for img in images if 'src' in img.attrs]

    def get_value(label):
        item = soup.find("span", string=label)
        if item:
            value = item.find_parent("dt").find_next_sibling("dd")
            if value:
                return value.get_text(strip=True)
        return "Не найдено"

    # Характеристики которые нахадяться на главной странице
    series = get_value("Серія")
    screen_size = get_value("Діагональ екрана")
    resolution = get_value("Роздільна здатність дисплея")
    
    # Функция для полного списока характеристик
    def get_characteristics(url_specs, headers):
        
        
        response = requests.get(url_specs, headers=headers)
        if response.status_code != 200:
            print("Ошибка запроса", response.status_code)
            return None
        
        soup = BeautifulSoup(response.text, "html.parser")
        # Создаем пустой словарь для хранения характеристик
        characteristics = {} 

        sections = soup.find_all("section", class_="group")
        for section in sections:
            heading = section.find("h3", class_="sub-heading")
            if heading:
                category = heading.text.strip()
                characteristics[category] = {}
                
                items = section.find_all("div", class_="item")
                for item in items:
                    label_elem = item.find("dt", class_="label")
                    if not label_elem:
                        continue  # Пропускаем, если label отсутствует
                    
                    label = label_elem.text.strip()
                    values = [v.text.strip() for v in item.find_all("span") if v.text.strip()]
                    
                    characteristics[category][label] = list(set(values))    

        return characteristics
    
    url_specs = url + "characteristics/"  #Добавляем раздел с характ
    specs = get_characteristics(url_specs, headers)
    
    # Вывод того что взяли с главной страницы
    print(f'Полное название товара: {product_name}')
    print(f'Цвет: {color}')
    print(f'Объем памяти: {storage}')
    print(f'Продавец: {seller}')
    print(f'Цена обычная: {price_normal}')
    print(f'Цена акционная: {price_promo}')
    print('Ссылки на фото товара:')
    for link in img_urls:
        print(f'  {link}')
    print(f'Код товара: {product_code}')
    print(f'Количество отзывов: {reviews_count}')
    print(f'Серия: {series}')
    print(f'Диагональ экрана: {screen_size}')
    print(f'Разрешение дисплея: {resolution}')
    print(f'Характеристики: {specs}')


    # Загружаем шаблон Excel
    wb = load_workbook("template.xlsx")
    ws = wb.active  # Берем первый лист

    # Заполняем данные
    ws["A2"] = product_name       # Название
    ws["B2"] = color              # Цвет
    ws["C2"] = storage            # Память
    ws["D2"] = seller             # Продавец
    ws["E2"] = price_normal       # Обычная цена
    ws["F2"] = price_promo        # Акционная цена
    ws["G2"] = product_code       # Код товара
    ws["H2"] = reviews_count      # Количество отзывов
    ws["I2"] = series             # Серия
    ws["J2"] = screen_size        # Диагональ
    ws["K2"] = resolution         # Разрешение
    for i, url in enumerate(img_urls, start=2):  # начинаем с 2, чтобы записывать с L2
        ws[f"L{i}"] = url         # Ссылки на фото
    # Создаем новый лист для характеристик, что бы было удобно читать
    ws_specs = wb.create_sheet("Характеристики")  # Новый лист
    row = 1
    if specs:
        for category, items in specs.items():
            ws_specs[f"A{row}"] = category  # Категория
            row += 1
            for key, value in items.items():
                ws_specs[f"B{row}"] = key      # Название характеристики
                ws_specs[f"C{row}"] = ", ".join(value)  # Значения характеристики
                row += 1
    # Сохраняем в новый файл
    wb.save("result.xlsx")
    print("✅ Данные успешно сохранены в result.xlsx")
else:
    print(f"Ошибка: {response.status_code}")
