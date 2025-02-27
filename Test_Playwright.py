from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# URL страницы товара
url = 'https://rozetka.com.ua/ua/apple-iphone-15-128gb-black/p395460480/'

def main():
    url = 'https://rozetka.com.ua/ua/apple-iphone-15-128gb-black/p395460480/'
    with sync_playwright() as p:
        # Запуск браузера
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        # Открываем страницу
        page.goto(url)

        # Ожидаем загрузки страницы
        page.wait_for_selector("h1")

        try:
            # Имя продукта
            product_name = page.query_selector("h1.title__font").inner_text().strip()

            # Цвет
            color = "Не найден"
            color_elements = page.query_selector_all('p.text-base.mb-2')
            for element in color_elements:
                if "Колір" in element.inner_text():
                    color = element.query_selector("span.bold").inner_text().strip()
                    break

            # Встроенная память
            storage = "Не найден"
            for element in color_elements:
                if "Вбудована пам'ять" in element.inner_text():
                    storage = element.query_selector("span.bold").inner_text().strip()
                    break

            # Продавец (логотип)
            seller = "Продавец не найден"
            try:
                seller_element = page.query_selector("span.seller-logo img")
                seller = seller_element.get_attribute("src") if seller_element else "Продавец не найден"
            except:
                seller = "Продавец не найден"

            # Цены
            price_promo = page.query_selector("p.product-price__big.product-price__big-color-red")
            price_promo = price_promo.inner_text().strip() if price_promo else "Нет акционной цены"

            price_normal = page.query_selector("p.product-price__small")
            price_normal = price_normal.inner_text().strip() if price_normal else "Нет обычной цены"

            # Извлечение кода товара
            code_element = page.query_selector('span.ms-auto.color-black-60')
            product_code = code_element.inner_text().strip().replace('Код: ', '') if code_element else 'Не найдено'

            # Количество отзывов
            reviews_count = page.query_selector("span.tabs__link-text")
            reviews_count = reviews_count.inner_text().strip() if reviews_count else "Нет отзывов"

            # Ссылки на изображения
            img_urls = []
            images = page.query_selector_all("ul.simple-slider__list img.image")
            for img in images:
                src = img.get_attribute("src")
                if src:
                    img_urls.append(src)

            # Функция для парсинга характеристик
            def get_characteristics():
                characteristics = {}

                # Переход на вкладку характеристик
                page.goto(url + "characteristics/")
                page.wait_for_selector("h1")

                sections = page.query_selector_all("section.group")
                for section in sections:
                    try:
                        category = section.query_selector("h3.sub-heading").inner_text().strip()
                        characteristics[category] = {}

                        items = section.query_selector_all("div.item")
                        for item in items:
                            try:
                                label = item.query_selector("dt.label").inner_text().strip()
                                values = [v.inner_text().strip() for v in item.query_selector_all("span") if v.inner_text().strip()]
                                characteristics[category][label] = ", ".join(values)
                            except Exception as ex:
                                print(f"Ошибка при обработке характеристики: {ex}")
                                continue
                    except Exception as ex:
                        print(f"Ошибка при обработке секции характеристик: {ex}")
                        continue
                return characteristics

            # Получаем характеристики
            specs = get_characteristics()

            # Извлечение серии
            series_element = page.query_selector('dt:has-text("Серія") + dd .button--link')
            series = series_element.inner_text().strip() if series_element else 'Не найдено'

            # Извлечение диагонали экрана
            screen_size_element = page.query_selector('dt:has-text("Діагональ екрана") + dd .button--link')
            screen_size = screen_size_element.inner_text().strip() if screen_size_element else 'Не найдено'

            # Извлечение разрешения дисплея
            resolution_element = page.query_selector('dt:has-text("Роздільна здатність дисплея") + dd .button--link')
            resolution = resolution_element.inner_text().strip() if resolution_element else 'Не найдено'


            # Вывод данных
            print(f'Полное название товара: {product_name}')
            print(f'Цвет: {color}')
            print(f'Объем памяти: {storage}')
            print(f'Продавец (логотип): {seller}')
            print(f'Цена обычная: {price_normal}')
            print(f'Цена акционная: {price_promo}')
            print('Ссылки на фото товара:')
            for link in img_urls:
                print(f'  {link}')
            print(f'Код товара: {product_code}')
            print(f'Количество отзывов: {reviews_count}')
            print(f'Серия: {series}')
            print(f'Диагональ экрана: {screen_size}')
            print(f'Разрешение дисплея: {resolution}')
            print(f'Характеристики: {specs}')

            # Запись в Excel
            wb = load_workbook("template.xlsx")
            ws = wb.active

            ws["A2"] = product_name
            ws["B2"] = color
            ws["C2"] = storage
            ws["D2"] = seller
            ws["E2"] = price_normal
            ws["F2"] = price_promo
            ws["G2"] = product_code
            ws["H2"] = reviews_count
            ws["I2"] = series
            ws["J2"] = screen_size
            ws["K2"] = resolution

            for i, url in enumerate(img_urls, start=2):
                ws[f"L{i}"] = url

            # Запись характеристик на отдельный лист
            ws_specs = wb.create_sheet("Характеристики")
            row = 1
            for category, items in specs.items():
                ws_specs[f"A{row}"] = category  # Категория
                row += 1
                for key, value in items.items():
                    ws_specs[f"B{row}"] = key  # Название характеристики
                    ws_specs[f"C{row}"] = value  # Значения характеристики
                    row += 1

            wb.save("result.xlsx")
            print("✅ Данные успешно сохранены в result.xlsx")

        except Exception as e:
            print("❌ Ошибка:", e)

        finally:
            browser.close()  # Закрываем браузер

if __name__ == "__main__":
    main()

